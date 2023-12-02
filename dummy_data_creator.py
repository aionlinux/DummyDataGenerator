import openai
import random
from faker import Faker
from fpdf import FPDF
from docx import Document
from openpyxl import Workbook
import os

# Initialize Faker
fake = Faker()

def generate_timeshare_data(api_key, complete=True):
    print("Generating timeshare data...")
    # Set the OpenAI API key using a secure method
    openai.api_key = api_key

    # Generate the basic fake data
    data = {
        "client_name": fake.name(),
        "email": fake.email(),
        "phone": fake.phone_number(),
        "location": fake.city(),
        "sales_rep": fake.name(),
        "interest_level": random.choice(["High", "Medium", "Low"]),
    }

    # Generate realistic notes using OpenAI GPT-3.5 Turbo, if complete
    if complete or random.random() > 0.1:  # 90% chance to include notes
        try:
            print("Calling OpenAI API for generating notes...")
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[{
                    "role": "system", 
                    "content": ("Write a concise note summarizing a customer's response in a timeshare sales "
                                "interaction, such as their interest level, any specific requests they made, "
                                "or intentions they expressed about the purchase.")
                }],
                max_tokens=50)
            data["notes"] = response['choices'][0]['message']['content'].strip()
            print("Notes generated successfully.")
        except openai.error.OpenAIError as e:
            print(f"An error occurred: {e}")
            data["notes"] = "Error generating notes."
    else:
        data["notes"] = ""
        print("No notes generated.")

    return data

def create_pdf(data, filename):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    for key, value in data.items():
        pdf.cell(200, 10, txt=f"{key}: {value}", ln=True)
    pdf.output(filename)
    print(f"PDF document '{filename}' created.")

def create_word(data, filename):
    doc = Document()
    for key, value in data.items():
        doc.add_paragraph(f"{key}: {value}")
    doc.save(filename)
    print(f"Word document '{filename}' created.")

def create_excel(data, filename):
    workbook = Workbook()
    sheet = workbook.active
    for i, (key, value) in enumerate(data.items(), start=1):
        sheet.cell(row=i, column=1, value=key)
        sheet.cell(row=i, column=2, value=value)
    workbook.save(filename)
    print(f"Excel document '{filename}' created.")

def generate_documents(num_docs, output_directory):
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("No OPENAI_API_KEY found in the environment variables.")
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
        print(f"Created output directory: {output_directory}")

    for i in range(num_docs):
        print(f"Generating document {i+1}/{num_docs}...")
        complete = random.random() > 0.05  # 95% chance to be complete
        data = generate_timeshare_data(api_key, complete)
        doc_type = random.choice(["pdf", "docx", "xlsx"])
        filename = os.path.join(output_directory, f"timeshare_{i}.{doc_type}")
        print(f"Creating a {doc_type} document...")
        if doc_type == "pdf":
            create_pdf(data, filename)
        elif doc_type == "docx":
            create_word(data, filename)
        elif doc_type == "xlsx":
            create_excel(data, filename)
        print(f"Document {filename} generated successfully.")

    print("All documents generated.")

# Example usage
output_directory = "/PATH/TO/YOUR/DESIRED/OUTPUT/DIRECTORY"
generate_documents(10, output_directory)
